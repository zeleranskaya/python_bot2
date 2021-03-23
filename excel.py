from openpyxl import load_workbook

book = load_workbook('таблица_бота.xlsx')

sheet_2 = book['листоу']
stickers_sheet = book['стикс']


print(book.worksheets)

print(sheet_2['A1'].value)

for i in range(1, 4):
    print(stickers_sheet.cell(row=1, column=i).value)

for a in 'ABC':
    print(stickers_sheet[f'{a}1'].value)
